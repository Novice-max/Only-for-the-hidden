from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook


def load_school_data(path: str):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")
    # keep existing behavior simple: delegate to pandas if available
    try:
        import pandas as pd

        return pd.read_excel(p)
    except Exception:
        # Fallback: return the openpyxl workbook object
        return load_workbook(p)


def _normalize(s: str) -> str:
    return "" if s is None else "".join(c for c in s.lower() if c.isalnum())


def _find_header_map(ws):
    """Return map of normalized header -> column index (1-based) from first non-empty row."""
    if ws.max_row < 1:
        return {}
    headers = [cell.value for cell in ws[1]]
    header_map = {}
    for idx, h in enumerate(headers, start=1):
        key = _normalize(h)
        if key:
            header_map[key] = idx
    return header_map


def _get_col_index_by_candidates(header_map: Dict[str, int], candidates: List[str]):
    for c in candidates:
        k = _normalize(c)
        if k in header_map:
            return header_map[k]
    return None


def apply_payment_to_excel(
    workbook_path: str,
    tx_id: str,
    amount: int,
    reference_order: List[str],
    allocations: Dict[str, int],
    remaining_credit: int,
    term: str,
) -> None:
    """Persist allocation results into the Excel workbook at `workbook_path`.

    This updates STUDENTS_MASTER (PaidTotal, Credit, Balance, Status), and appends
    rows to TRANSACTIONS, ALLOCATIONS and CREDITS (when applicable).

    Raises ValueError for missing expected sheets/columns or invalid inputs.
    """
    if not isinstance(amount, int) or amount < 0:
        raise ValueError("amount must be a non-negative int")
    if not isinstance(remaining_credit, int) or remaining_credit < 0:
        raise ValueError("remaining_credit must be a non-negative int")
    if not isinstance(reference_order, list) or not all(isinstance(r, str) for r in reference_order):
        raise ValueError("reference_order must be a list of strings")
    if not isinstance(allocations, dict):
        raise ValueError("allocations must be a dict mapping admission_no to int")

    wb_path = Path(workbook_path)
    if not wb_path.exists():
        raise FileNotFoundError(f"Excel file not found: {workbook_path}")

    wb = load_workbook(wb_path)

    # Load STUDENTS_MASTER
    if "STUDENTS_MASTER" not in wb.sheetnames:
        raise ValueError("STUDENTS_MASTER sheet not found in workbook")
    ws_students = wb["STUDENTS_MASTER"]
    header_map = _find_header_map(ws_students)

    adm_col = _get_col_index_by_candidates(header_map, ["admission_no", "admissionno", "admission", "admissionnumber", "admissionno"])
    paid_col = _get_col_index_by_candidates(header_map, ["paidtotal", "paid_total", "paidtotal"])
    credit_col = _get_col_index_by_candidates(header_map, ["credit"])
    balance_col = _get_col_index_by_candidates(header_map, ["balance"])
    status_col = _get_col_index_by_candidates(header_map, ["status"])

    if adm_col is None or paid_col is None or credit_col is None or balance_col is None or status_col is None:
        raise ValueError("Required columns (admission_no, PaidTotal, Credit, Balance, Status) not all present in STUDENTS_MASTER")

    # Build admission_no -> row index map
    adm_row_map = {}
    for row in range(2, ws_students.max_row + 1):
        cell = ws_students.cell(row=row, column=adm_col).value
        if cell is None:
            continue
        adm_row_map[str(cell).strip()] = row

    # Update students based on allocations
    # Keep track of credit increments to apply later if remaining_credit split
    credit_increments: Dict[str, int] = {}

    for adm, alloc_amt in allocations.items():
        if not isinstance(adm, str):
            raise ValueError("allocation keys must be admission numbers (strings)")
        if not isinstance(alloc_amt, int) or alloc_amt < 0:
            raise ValueError("allocation amounts must be non-negative ints")
        row = adm_row_map.get(adm)
        if row is None:
            raise ValueError(f"admission_no '{adm}' not found in STUDENTS_MASTER")

        # Read existing values
        old_paid = ws_students.cell(row=row, column=paid_col).value or 0
        old_credit = ws_students.cell(row=row, column=credit_col).value or 0
        old_balance = ws_students.cell(row=row, column=balance_col).value or 0

        # Ensure numeric types
        try:
            old_paid = int(old_paid)
        except Exception:
            old_paid = 0
        try:
            old_credit = int(old_credit)
        except Exception:
            old_credit = 0
        try:
            old_balance = int(old_balance)
        except Exception:
            old_balance = 0

        new_paid = old_paid + alloc_amt
        # Allocation reduces the outstanding balance
        new_balance = old_balance - alloc_amt

        ws_students.cell(row=row, column=paid_col, value=new_paid)
        ws_students.cell(row=row, column=balance_col, value=new_balance)

        # Temporarily leave credit unchanged; credits from remaining_credit handled below
        # Set status based on new_balance (PARTIAL / PAID / OVERPAID)
        if new_balance > 0:
            status_val = "PARTIAL"
        elif new_balance == 0:
            status_val = "PAID"
        else:
            status_val = "OVERPAID"
        ws_students.cell(row=row, column=status_col, value=status_val)

    # Handle remaining_credit splitting across reference_order if any
    credits_appended: Dict[str, int] = {}
    if remaining_credit > 0 and reference_order:
        refs = [r for r in reference_order if r in adm_row_map]
        if refs:
            n = len(refs)
            base = remaining_credit // n
            rem = remaining_credit % n
            for idx, adm in enumerate(refs):
                share = base + (1 if idx < rem else 0)
                if share <= 0:
                    continue
                row = adm_row_map[adm]
                old_credit = ws_students.cell(row=row, column=credit_col).value or 0
                try:
                    old_credit = int(old_credit)
                except Exception:
                    old_credit = 0
                new_credit = old_credit + share
                ws_students.cell(row=row, column=credit_col, value=new_credit)

                # Adjust balance as credits reduce outstanding
                old_balance = ws_students.cell(row=row, column=balance_col).value or 0
                try:
                    old_balance = int(old_balance)
                except Exception:
                    old_balance = 0
                new_balance = old_balance - share
                ws_students.cell(row=row, column=balance_col, value=new_balance)

                # Update status based on new_balance (PARTIAL / PAID / OVERPAID)
                if new_balance > 0:
                    status_val = "PARTIAL"
                elif new_balance == 0:
                    status_val = "PAID"
                else:
                    status_val = "OVERPAID"
                ws_students.cell(row=row, column=status_col, value=status_val)

                credits_appended[adm] = share

    # Append to TRANSACTIONS sheet
    tx_sheet_name = "TRANSACTIONS"
    if tx_sheet_name not in wb.sheetnames:
        ws_tx = wb.create_sheet(tx_sheet_name)
        ws_tx.append(["TxID", "Amount", "Term", "ReferenceOrder", "RemainingCredit"])
    else:
        ws_tx = wb[tx_sheet_name]

    ref_joined = "|".join(reference_order)
    ws_tx.append([tx_id, amount, term, ref_joined, remaining_credit])

    # Append to ALLOCATIONS sheet (one row per allocation)
    alloc_sheet_name = "ALLOCATIONS"
    if alloc_sheet_name not in wb.sheetnames:
        ws_alloc = wb.create_sheet(alloc_sheet_name)
        ws_alloc.append(["TxID", "AdmissionNo", "AllocatedAmount"])
    else:
        ws_alloc = wb[alloc_sheet_name]

    for adm, alloc_amt in allocations.items():
        ws_alloc.append([tx_id, adm, alloc_amt])

    # Append to CREDITS sheet if any credits were applied or remaining_credit>0
    credit_sheet_name = "CREDITS"
    if credit_sheet_name not in wb.sheetnames:
        ws_cred = wb.create_sheet(credit_sheet_name)
        ws_cred.append(["TxID", "AdmissionNo", "CreditAmount"])
    else:
        ws_cred = wb[credit_sheet_name]

    for adm, share in credits_appended.items():
        ws_cred.append([tx_id, adm, share])

    # Save workbook (open -> modify -> save)
    wb.save(wb_path)
