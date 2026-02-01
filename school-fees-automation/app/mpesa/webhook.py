from fastapi import APIRouter, Request, HTTPException
from typing import Any, Dict

from app.mpesa.parser import parse_reference
from app.core.allocator import allocate_payment
from app.core.excel import _find_header_map, _normalize, apply_payment_to_excel
from openpyxl import load_workbook

router = APIRouter()

# Default locations; can be changed if you wire config
EXCEL_PATH = "data/SCHOOL_FEES_AUTOMATION.xlsx"
CURRENT_TERM = "2026-T1"


def _find_students_from_workbook(path: str):
    wb = load_workbook(path, data_only=True)
    if "STUDENTS_MASTER" not in wb.sheetnames:
        raise ValueError("STUDENTS_MASTER sheet not found in workbook")
    ws = wb["STUDENTS_MASTER"]
    header_map = _find_header_map(ws)
    adm_col = None
    bal_col = None
    # Find admission and balance columns (reuse normalization)
    for cand in ["admission_no", "admissionno", "admission", "admissionnumber"]:
        idx = _normalize(cand)
        if idx and idx in header_map:
            adm_col = header_map[idx]
            break
    for cand in ["balance"]:
        idx = _normalize(cand)
        if idx and idx in header_map:
            bal_col = header_map[idx]
            break
    if adm_col is None or bal_col is None:
        raise ValueError("Required columns (admission_no, Balance) not found in STUDENTS_MASTER")

    students = []
    for row in range(2, ws.max_row + 1):
        adm_cell = ws.cell(row=row, column=adm_col).value
        if adm_cell is None:
            continue
        adm = str(adm_cell).strip()
        bal_cell = ws.cell(row=row, column=bal_col).value or 0
        try:
            bal = int(bal_cell)
        except Exception:
            bal = 0
        students.append({"admission_no": adm, "balance": bal})
    wb.close()
    return students


@router.post("/callback")
async def mpesa_callback(request: Request) -> Dict[str, Any]:
    payload = await request.json()

    # Validate required fields
    tx_id = payload.get("TransID")
    if not isinstance(tx_id, str) or not tx_id:
        raise HTTPException(status_code=400, detail="TransID is required and must be a string")

    trans_amount = payload.get("TransAmount")
    if not isinstance(trans_amount, int):
        # Accept numeric but enforce int
        if isinstance(trans_amount, float) and trans_amount.is_integer():
            trans_amount = int(trans_amount)
        else:
            raise HTTPException(status_code=400, detail="TransAmount is required and must be an integer amount")

    raw_ref = payload.get("BillRefNumber")
    if not isinstance(raw_ref, str) or not raw_ref:
        raise HTTPException(status_code=400, detail="BillRefNumber is required and must be a string")

    # Parse reference order
    try:
        reference_order = parse_reference(raw_ref)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # Load students read-only from Excel
    try:
        students = _find_students_from_workbook(EXCEL_PATH)
    except Exception:
        # Let FastAPI convert to 500 for unexpected errors
        raise

    # Allocate
    result = allocate_payment(students, trans_amount, reference_order)

    # Persist to Excel (may raise on failure)
    apply_payment_to_excel(
        workbook_path=EXCEL_PATH,
        tx_id=tx_id,
        amount=trans_amount,
        reference_order=reference_order,
        allocations=result.get("allocations", {}),
        remaining_credit=result.get("remaining_credit", 0),
        term=CURRENT_TERM,
    )

    return {"status": "ok"}
